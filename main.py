import os
from openpyxl import load_workbook
import smtplib
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


def create_dir(dir_path):
    """
    Функция создает директорию по указаннному пути
    :param dir_path:
    :return:
    """

    dir = os.path.join(os.path.dirname(__file__), dir_path)
    try:
        os.stat(dir)
    except FileNotFoundError:
        os.mkdir(dir)


def create_file(file_path, text=None):
    """
    Функция создает файл и записывает в него переданный текст
    :param file_path:
    :param text:
    :return:
    """
    file = os.path.join(os.path.dirname(__file__), file_path)
    try:
        os.stat(file)
    except FileNotFoundError:
        with open(file, "w") as f:
            if text:
                f.write("Hello, world")


def read_excel_to_file(excel_file, file_to_write):
    """
    Функция читает файл Excel, выбирает нужные поля и записывает данные в файл
    :param excel_file:
    :param file_to_write:
    :return:
    """

    # Открываем файл Excel
    try:
        wb = load_workbook(filename = excel_file)
    except FileNotFoundError:
        print("Неверный путь к файлу Excel")
        return

    # Получаем имя первого листа рабочей книги
    first_sheet_name = wb.sheetnames[0]
    # sheet_ranges = wb['Лист1']
    sheet = wb[first_sheet_name]

    # Получаем количесво строк
    max_row = sheet.max_row
    # Для каждой строки
    for i in range(1, max_row + 1):
        # Получаем код
        code = sheet['C' + str(i)].value
        # Получаем значение
        value = str(sheet['E' + str(i)].value)
        # Пропускаем пустые строки, заголовки и строки со значениями, котрые невозможно привести к числу
        if not value or not value.isdigit():
            continue
        # Записываем строку в файл
        str_to_write = 'Kate wrote {value} и eat {code}\n'.format(value=value,code=code)
        # print(str_to_write)
        with open(file_to_write, 'a+') as f:
            f.write(str_to_write)


def send_mail(host, username, password, mail_from, mail_to, subject, text, port=25, ssl=0):
    """
    Метод send отправляет сообщение
    :param host: str
    :param username: str
    :param password_hash: str
    :param mail_from: str
    :param mail_to: str
    :param subject: str
    :param text: str
    :param port: int
    :param tls: int
    :return: None
    """

    # Подготавливаем сообщение
    msg = MIMEMultipart()
    msg["Subject"] = Header(subject, 'utf-8')
    msg["From"] = Header(mail_from, 'utf-8')
    msg["To"] = Header(mail_to, 'utf-8')
    text = MIMEText(('<br><h3>' + text + '</h3>').encode('utf-8'), 'html', _charset='utf-8')
    msg.attach(text)

    if ssl:
        server = smtplib.SMTP_SSL(host, port)
    else:
        server = smtplib.SMTP(host, port)
    server.login(username, password)

    # Отправляем сообщение
    server.sendmail(mail_from, mail_to, msg.as_string())
    server.quit()


if __name__ == '__main__':

    # Создаем директорию
    create_dir('[A]')
    # Создаем необходимые файлы и записываем в них нужный текст
    create_file('File1.txt', "Hello, world")
    create_file('File2.txt', "Hello, world")
    # Читаем файл Excel и записываем данные в указанный файл
    read_excel_to_file('file.xlsx', '[A]/File3.txt')


###### Отправляем почту
    # Настройки SMTP - тестовый аккаунт - можно экпериментировать
    username = 'mymail@yandex.ru'
    password = '1111111111'
    host = 'smtp.yandex.ru'
    port = 465
    # от кого
    mail_from = 'mymail-2019@yandex.ru'

   ############ Сообщение 1 - пункт 7 (непонятно - что такое получившееся имя файла?)

    # кому
    mail_to = 'vasya@yandex.ru'
    # Тема
    subject = 'Hello, file3'
    # текст сообщения
    text = """ Hello world"""

    send_mail(host, username, password, mail_from, mail_to, subject, text, port, 1)

    ############# Сообщение 2 - пункт 8

    # кому
    mail_to = 'vasya@yandex.ru'
    # Тема
    subject = '[Предмет]'
    # текст сообщения
    text = """ Здесь сам текст """

    send_mail(host, username, password, mail_from, mail_to, subject, text, port, 1)

    #################### Сообщение 2 - пункт 8

    # кому
    mail_to = 'vasya@yandex.ru'
    # Тема
    subject = '[Предмет]'
    # текст сообщения
    text = """ Здесь сам текст  """

    send_mail(host, username, password, mail_from, mail_to, subject, text, port, 1)

